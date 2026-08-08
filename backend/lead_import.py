import io
import re
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from activities import activities_for_new_lead, record_activities
from database import Lead, LeadImportBatch
from lead_attachments import delete_attachments_for_lead
from lead_automation import apply_lead_automation

MAX_IMPORT_ROWS = 500

IMPORT_FIELDS: list[tuple[str, str, bool]] = [
    ("isletme_adi", "İşletme Adı", True),
    ("yetkili", "Yetkili", False),
    ("sehir", "Şehir", False),
    ("instagram", "Instagram", False),
    ("whatsapp", "WhatsApp / Telefon", False),
    ("eposta", "E-posta", False),
    ("ilk_iletisim_kanali", "İlk İletişim Kanalı", False),
    ("ilk_mesaj_tarihi", "İlk Mesaj Tarihi", False),
    ("ilk_mesaj_saati", "İlk Mesaj Saati", False),
    ("durum", "Durum", False),
    ("oncelik", "Öncelik", False),
    ("takip_1", "Takip 1", False),
    ("takip_2", "Takip 2", False),
    ("demo_gonderildi", "Demo Gönderildi", False),
    ("demo_tarihi", "Demo Tarihi", False),
    ("gorusme_tarihi", "Görüşme Tarihi", False),
    ("gorusme_saati", "Görüşme Saati", False),
    ("teklif", "Teklif", False),
    ("sonuc", "Sonuç", False),
    ("satis_tutari", "Satış Tutarı", False),
    ("satis_tarihi", "Satış Tarihi", False),
    ("notlar", "Notlar", False),
]

HEADER_TO_FIELD = {label.lower(): key for key, label, _ in IMPORT_FIELDS}
HEADER_TO_FIELD.update({key: key for key, _, _ in IMPORT_FIELDS})

VALID_PRIORITIES = {"dusuk", "orta", "yuksek"}
PRIORITY_ALIASES = {
    "düşük": "dusuk",
    "dusuk": "dusuk",
    "low": "dusuk",
    "orta": "orta",
    "medium": "orta",
    "yüksek": "yuksek",
    "yuksek": "yuksek",
    "high": "yuksek",
}

KNOWN_DURUM = {
    "yeni",
    "iletişime geçildi",
    "takip bekliyor",
    "demo gönderildi",
    "cevap yok",
    "görüşme planlandı",
    "teklif verildi",
    "müşteri",
    "red",
    "kayıp",
}

DEMO_TRUE = {"evet", "yes", "true", "1", "x", "✓"}


def build_import_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Musteriler"

    headers = [f"{label} *" if required else label for _, label, required in IMPORT_FIELDS]
    ws.append(headers)
    ws.append([
        "Örnek Kuaför Salonu",
        "Ahmet Yılmaz",
        "İstanbul",
        "@ornekkuaför",
        "05551234567",
        "info@ornek.com",
        "Instagram DM",
        "2026-07-17",
        "10:30",
        "Yeni",
        "orta",
        "",
        "",
        "Hayır",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "Excel ile toplu aktarım örneği",
    ])

    for col_idx, (_, label, _) in enumerate(IMPORT_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(24, len(label) + 4))

    info = wb.create_sheet("Aciklama")
    info.append(["BehTech Sales Hub — Toplu müşteri içe aktarma"])
    info.append([])
    info.append(["• İşletme Adı zorunludur (*)."])
    info.append(["• Boş satırlar atlanır."])
    info.append(["• Tarih formatı: YYYY-MM-DD (ör. 2026-07-17)"])
    info.append(["• Öncelik: dusuk, orta, yuksek"])
    info.append(["• Demo Gönderildi: Evet veya Hayır"])
    info.append(["• İçe aktarmadan önce panelden kategori seçin (ör. Müşteriler)."])
    info.append([f"• En fazla {MAX_IMPORT_ROWS} satır yüklenebilir."])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Evet" if value else "Hayır"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M") if value.time().replace(microsecond=0) != datetime.min.time() else value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_bool(value: str) -> bool:
    return value.strip().lower() in DEMO_TRUE


def _parse_priority(value: str) -> str:
    raw = value.strip().lower()
    if not raw:
        return "orta"
    return PRIORITY_ALIASES.get(raw, raw)


def _parse_durum(value: str) -> str:
    raw = value.strip()
    if not raw:
        return "Yeni"
    if raw.lower() in KNOWN_DURUM:
        for status in (
            "Yeni",
            "İletişime Geçildi",
            "Takip Bekliyor",
            "Demo Gönderildi",
            "Cevap Yok",
            "Görüşme Planlandı",
            "Teklif Verildi",
            "Müşteri",
            "Red",
            "Kayıp",
        ):
            if status.lower() == raw.lower():
                return status
    return raw


def _parse_amount(value: str) -> float:
    if not value.strip():
        return 0.0
    cleaned = re.sub(r"[₺\s]", "", value.strip())
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    return float(cleaned)


def _map_headers(header_row: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, header in enumerate(header_row):
        normalized = _cell_text(header).lower().replace("*", "").strip()
        key = HEADER_TO_FIELD.get(normalized)
        if key:
            mapping[idx] = key
    return mapping


def _row_is_empty(values: dict[str, str]) -> bool:
    return not any(v.strip() for v in values.values())


def parse_leads_from_xlsx(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("Excel dosyası boş") from exc

    column_map = _map_headers(list(header_row))
    if "isletme_adi" not in column_map.values():
        raise ValueError('Excel dosyasında "İşletme Adı" sütunu bulunamadı. Şablonu indirip kullanın.')

    parsed: list[dict[str, str]] = []
    for row in rows:
        values: dict[str, str] = {key: "" for key, _, _ in IMPORT_FIELDS}
        for col_idx, field in column_map.items():
            if col_idx < len(row):
                values[field] = _cell_text(row[col_idx])
        if _row_is_empty(values):
            continue
        parsed.append(values)
        if len(parsed) > MAX_IMPORT_ROWS:
            raise ValueError(f"En fazla {MAX_IMPORT_ROWS} satır içe aktarılabilir")
    return parsed


def import_leads_from_rows(
    db: Session,
    *,
    user_id: int,
    category: str,
    rows: list[dict[str, str]],
    filename: str = "",
) -> dict[str, Any]:
    created = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    batch = LeadImportBatch(
        user_id=user_id,
        category=category,
        filename=filename or "",
    )
    db.add(batch)
    db.flush()

    for index, row in enumerate(rows, start=2):
        isletme_adi = row.get("isletme_adi", "").strip()
        if not isletme_adi:
            skipped += 1
            errors.append({"row": index, "isletme_adi": "", "error": "İşletme adı boş"})
            continue

        oncelik = _parse_priority(row.get("oncelik", ""))
        if oncelik not in VALID_PRIORITIES:
            errors.append(
                {
                    "row": index,
                    "isletme_adi": isletme_adi,
                    "error": "Geçersiz öncelik (dusuk, orta, yuksek)",
                }
            )
            continue

        try:
            satis_tutari = _parse_amount(row.get("satis_tutari", ""))
        except ValueError:
            errors.append(
                {"row": index, "isletme_adi": isletme_adi, "error": "Satış tutarı geçersiz"}
            )
            continue

        payload = {
            "isletme_adi": isletme_adi,
            "yetkili": row.get("yetkili", "").strip(),
            "sehir": row.get("sehir", "").strip(),
            "instagram": row.get("instagram", "").strip(),
            "whatsapp": row.get("whatsapp", "").strip(),
            "eposta": row.get("eposta", "").strip(),
            "ilk_iletisim_kanali": row.get("ilk_iletisim_kanali", "").strip(),
            "ilk_mesaj_tarihi": row.get("ilk_mesaj_tarihi", "").strip()[:10],
            "ilk_mesaj_saati": row.get("ilk_mesaj_saati", "").strip()[:5],
            "durum": _parse_durum(row.get("durum", "")),
            "oncelik": oncelik,
            "takip_1": row.get("takip_1", "").strip(),
            "takip_2": row.get("takip_2", "").strip(),
            "demo_gonderildi": _parse_bool(row.get("demo_gonderildi", "")),
            "demo_tarihi": row.get("demo_tarihi", "").strip()[:10],
            "gorusme_tarihi": row.get("gorusme_tarihi", "").strip()[:10],
            "gorusme_saati": row.get("gorusme_saati", "").strip()[:5],
            "teklif": row.get("teklif", "").strip(),
            "sonuc": row.get("sonuc", "").strip(),
            "satis_tutari": satis_tutari,
            "satis_tarihi": row.get("satis_tarihi", "").strip()[:10],
            "notlar": row.get("notlar", "").strip(),
        }

        lead_fields = apply_lead_automation(payload)
        lead = Lead(
            user_id=user_id,
            category=category,
            import_batch_id=batch.id,
            **lead_fields,
        )
        db.add(lead)
        db.flush()
        record_activities(db, user_id, lead.id, activities_for_new_lead(payload))
        created += 1

    batch.created_count = created
    batch.failed_count = len(errors)
    batch.skipped_count = skipped

    if created:
        db.commit()
    else:
        db.rollback()

    return {
        "created": created,
        "failed": len(errors),
        "skipped": skipped,
        "batch_id": batch.id if created else None,
        "errors": errors[:50],
    }


def list_import_batches(db: Session, user_id: int) -> list[dict[str, Any]]:
    batches = (
        db.query(LeadImportBatch)
        .filter(LeadImportBatch.user_id == user_id)
        .order_by(LeadImportBatch.created_at.desc(), LeadImportBatch.id.desc())
        .all()
    )
    result: list[dict[str, Any]] = []
    for batch in batches:
        lead_count = (
            db.query(Lead)
            .filter(Lead.user_id == user_id, Lead.import_batch_id == batch.id)
            .count()
        )
        result.append(
            {
                "id": batch.id,
                "category": batch.category,
                "filename": batch.filename,
                "created_count": batch.created_count,
                "failed_count": batch.failed_count,
                "skipped_count": batch.skipped_count,
                "lead_count": lead_count,
                "created_at": batch.created_at,
            }
        )
    return result


def delete_import_batch(db: Session, user_id: int, batch_id: int) -> int:
    batch = (
        db.query(LeadImportBatch)
        .filter(LeadImportBatch.id == batch_id, LeadImportBatch.user_id == user_id)
        .first()
    )
    if not batch:
        raise LookupError("Import batch not found")

    leads = (
        db.query(Lead)
        .filter(Lead.user_id == user_id, Lead.import_batch_id == batch_id)
        .all()
    )
    for lead in leads:
        delete_attachments_for_lead(db, user_id, lead.id)
        db.delete(lead)

    deleted = len(leads)
    db.delete(batch)
    db.commit()
    return deleted

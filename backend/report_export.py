import csv
import io
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # type: ignore


def _currency(value: float | None) -> str:
    if value is None:
        return "—"
    return f"₺{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def export_report_csv(report: dict) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow(["BehTech Sales Hub Raporu"])
    writer.writerow(["Dönem", report["period_label"]])
    writer.writerow(["Başlangıç", report["period_start"]])
    writer.writerow(["Bitiş", report["period_end"]])
    writer.writerow([])

    writer.writerow(["Özet"])
    writer.writerow(["Yeni Kayıt", report["yeni_kayit"]])
    writer.writerow(["Yeni Müşteri", report["yeni_musteri"]])
    writer.writerow(["Dönüşüm Oranı (%)", report["donusum_orani"] if report["donusum_orani"] is not None else ""])
    if report.get("satis_sayisi") is not None:
        writer.writerow(["Satış Sayısı", report["satis_sayisi"]])
        writer.writerow(["Toplam Gelir", report["toplam_gelir"]])
        writer.writerow(["Ortalama Satış", report["ortalama_satis"]])
    writer.writerow([])

    writer.writerow(["Satış Hunisi"])
    writer.writerow(["Aşama", "Adet", "Dönüşüm (%)", "Genel (%)"])
    for stage in report["satis_hunisi"]:
        writer.writerow([
            stage["label"],
            stage["count"],
            stage["conversion_rate"] if stage["conversion_rate"] is not None else "",
            stage["overall_rate"] if stage["overall_rate"] is not None else "",
        ])
    writer.writerow([])

    writer.writerow(["Durum Dağılımı"])
    writer.writerow(["Durum", "Adet"])
    for item in report["durum_dagilimi"]:
        writer.writerow([item["durum"], item["count"]])
    writer.writerow([])

    writer.writerow(["Kategori Özeti"])
    writer.writerow(["Kategori", "Yeni Kayıt", "Müşteri"])
    for item in report["kategori_ozet"]:
        writer.writerow([item["category_label"], item["yeni_kayit"], item["musteri"]])

    if report.get("donem_satislar"):
        writer.writerow([])
        writer.writerow(["Dönem Satışları"])
        writer.writerow(["İşletme", "Kategori", "Şehir", "Tutar", "Tarih"])
        for sale in report["donem_satislar"]:
            writer.writerow([
                sale["isletme_adi"],
                sale["category_label"],
                sale["sehir"],
                sale["satis_tutari"],
                sale["satis_tarihi"],
            ])

    return buffer.getvalue().encode("utf-8-sig")


def export_report_xlsx(report: dict) -> bytes:
    if Workbook is None:
        raise RuntimeError("Excel dışa aktarma kullanılamıyor")

    wb = Workbook()
    ws = wb.active
    ws.title = "Özet"

    ws.append(["BehTech Sales Hub Raporu"])
    ws.append(["Dönem", report["period_label"]])
    ws.append(["Yeni Kayıt", report["yeni_kayit"]])
    ws.append(["Yeni Müşteri", report["yeni_musteri"]])
    ws.append(["Dönüşüm (%)", report["donusum_orani"]])
    if report.get("satis_sayisi") is not None:
        ws.append(["Satış Sayısı", report["satis_sayisi"]])
        ws.append(["Toplam Gelir", report["toplam_gelir"]])
        ws.append(["Ortalama Satış", report["ortalama_satis"]])

    funnel_ws = wb.create_sheet("Satış Hunisi")
    funnel_ws.append(["Aşama", "Adet", "Dönüşüm (%)", "Genel (%)"])
    for stage in report["satis_hunisi"]:
        funnel_ws.append([
            stage["label"],
            stage["count"],
            stage["conversion_rate"],
            stage["overall_rate"],
        ])

    status_ws = wb.create_sheet("Durumlar")
    status_ws.append(["Durum", "Adet"])
    for item in report["durum_dagilimi"]:
        status_ws.append([item["durum"], item["count"]])

    cat_ws = wb.create_sheet("Kategoriler")
    cat_ws.append(["Kategori", "Yeni Kayıt", "Müşteri"])
    for item in report["kategori_ozet"]:
        cat_ws.append([item["category_label"], item["yeni_kayit"], item["musteri"]])

    if report.get("donem_satislar"):
        sales_ws = wb.create_sheet("Satışlar")
        sales_ws.append(["İşletme", "Kategori", "Şehir", "Tutar", "Tarih"])
        for sale in report["donem_satislar"]:
            sales_ws.append([
                sale["isletme_adi"],
                sale["category_label"],
                sale["sehir"],
                sale["satis_tutari"],
                sale["satis_tarihi"],
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_report_pdf(report: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Heading1"], fontSize=18, spaceAfter=12)
    subtitle_style = ParagraphStyle("ReportSubtitle", parent=styles["Normal"], textColor=colors.grey, spaceAfter=16)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=13, spaceBefore=14, spaceAfter=8)

    period_title = "Haftalık Rapor" if report["period_type"] == "weekly" else "Aylık Rapor"
    story: list[Any] = [
        Paragraph(f"BehTech Sales Hub — {period_title}", title_style),
        Paragraph(report["period_label"], subtitle_style),
    ]

    summary_rows = [
        ["Yeni Kayıt", str(report["yeni_kayit"])],
        ["Yeni Müşteri", str(report["yeni_musteri"])],
        ["Dönüşüm Oranı", f"%{report['donusum_orani']}" if report["donusum_orani"] is not None else "—"],
    ]
    if report.get("satis_sayisi") is not None:
        summary_rows.extend([
            ["Satış Sayısı", str(report["satis_sayisi"])],
            ["Toplam Gelir", _currency(report["toplam_gelir"])],
            ["Ortalama Satış", _currency(report["ortalama_satis"])],
        ])

    story.append(Paragraph("Özet", section_style))
    summary_table = Table([["Metrik", "Değer"], *summary_rows], colWidths=[8 * cm, 8 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3432c7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafbff")]),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("Satış Hunisi", section_style))
    funnel_rows = [["Aşama", "Adet", "Dönüşüm", "Genel"]]
    for stage in report["satis_hunisi"]:
        funnel_rows.append([
            stage["label"],
            str(stage["count"]),
            f"%{stage['conversion_rate']}" if stage["conversion_rate"] is not None else "—",
            f"%{stage['overall_rate']}" if stage["overall_rate"] is not None else "—",
        ])
    funnel_table = Table(funnel_rows, colWidths=[5 * cm, 3 * cm, 4 * cm, 4 * cm])
    funnel_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3432c7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(funnel_table)

    if report["kategori_ozet"]:
        story.append(Paragraph("Kategori Özeti", section_style))
        cat_rows = [["Kategori", "Yeni Kayıt", "Müşteri"]]
        for item in report["kategori_ozet"][:10]:
            cat_rows.append([item["category_label"], str(item["yeni_kayit"]), str(item["musteri"])])
        cat_table = Table(cat_rows, colWidths=[8 * cm, 4 * cm, 4 * cm])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3432c7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(cat_table)

    doc.build(story)
    return buffer.getvalue()
